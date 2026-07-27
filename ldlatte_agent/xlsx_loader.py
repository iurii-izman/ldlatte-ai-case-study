from __future__ import annotations

import re
from pathlib import Path
from typing import BinaryIO
from urllib.parse import urlparse

from openpyxl import load_workbook

from .models import SeedProfile

INSTAGRAM_HOSTS = {"instagram.com", "www.instagram.com", "m.instagram.com"}
HANDLE_RE = re.compile(r"^[a-z0-9._]{1,30}$", re.IGNORECASE)


def normalize_instagram(value: str) -> tuple[str, str] | None:
    """Return (lowercase handle, canonical URL) or None.

    Handles are deliberately derived from the URL path, not from display text.
    Query parameters such as igsh/utm_source are discarded.
    """
    raw = (value or "").strip()
    if not raw:
        return None
    if raw.startswith("@"):
        handle = raw[1:].strip().lower()
    else:
        candidate = raw if "://" in raw else f"https://{raw}"
        parsed = urlparse(candidate)
        host = parsed.netloc.lower()
        if host not in INSTAGRAM_HOSTS:
            return None
        segments = [part for part in parsed.path.split("/") if part]
        if not segments:
            return None
        handle = segments[0].lower()
    if handle in {"p", "reel", "reels", "stories", "explore"}:
        return None
    if not HANDLE_RE.fullmatch(handle):
        return None
    return handle, f"https://www.instagram.com/{handle}/"


def load_seed_profiles(
    path_or_file: str | Path | BinaryIO,
    sheet_name: str = "Исходник",
) -> tuple[list[SeedProfile], dict[str, object]]:
    workbook = load_workbook(path_or_file, read_only=False, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    profiles: list[SeedProfile] = []
    seen: set[str] = set()
    hyperlink_overrides = 0
    skipped = 0

    for row in range(2, worksheet.max_row + 1):
        number_cell = worksheet.cell(row=row, column=1)
        profile_cell = worksheet.cell(row=row, column=2)
        display = str(profile_cell.value or "").strip()
        target = (
            str(profile_cell.hyperlink.target).strip()
            if profile_cell.hyperlink and profile_cell.hyperlink.target
            else display
        )
        if target and display and target != display:
            hyperlink_overrides += 1
        normalized = normalize_instagram(target)
        if normalized is None:
            if display or target:
                skipped += 1
            continue
        handle, normalized_url = normalized
        if handle in seen:
            continue
        seen.add(handle)
        number_value = number_cell.value
        number = (
            int(number_value)
            if isinstance(number_value, (int, float))
            and not isinstance(number_value, bool)
            and float(number_value).is_integer()
            else None
        )
        profiles.append(
            SeedProfile(
                excel_row=row,
                number=number,
                display=display,
                source_url=target,
                handle=handle,
                normalized_url=normalized_url,
            )
        )

    quality = {
        "sheet": worksheet.title,
        "unique_profiles": len(profiles),
        "hyperlink_overrides": hyperlink_overrides,
        "skipped_non_profiles": skipped,
        "duplicate_handles": 0,
        "note": (
            "Приоритет отдан hyperlink.target. Это исправляет скрытые расхождения "
            "между отображаемым текстом и настоящим адресом."
        ),
    }
    return profiles, quality
